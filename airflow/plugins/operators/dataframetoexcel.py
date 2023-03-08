from pathlib import Path
from typing import Any, Optional, Union, Mapping, Iterable, List

import pandas as pd
from airflow.models import BaseOperator
from openpyxl import load_workbook


class DataFrame2Excel(BaseOperator):
    """
    Converts pandas.DataFrame to xlsx file with limited file formatting options

    :param df: pandas.DataFrame object
    :type df: pandas.DataFrame: (templated)
    :param file_name: File name for data in df
    :type file_name: str (templated)
    :param xlsx_params: a dictionary of keyword arguments that will get passed
        to df.to_excel as well holds some additional key, values for additional
        file formatting. You can pass any key, value which is present int df.to_excel.

        **Example**: for xlsx_params dict object ::

            {
                'index': False,
                'group_to_sheet': "group_by_column_name",
                'sheet_name': "custom_sheet_name"
                'title': 'a title which will be placed in upper left cell above data table',
                'startcol': 0,
                'subtotals": {
                                'group_by': ['col1', 'col2'],
                                'index': ['index_col_name'],
                                'values': ['value_col1', 'value_col2', ..., 'value_coln'],
                                'aggfunc': 'sum',
                             }
            }

    :type xlsx_params: dict (templated)
    :param file_exists: if the file already exists; default is False
    :type file_exists: bool
    """

    PREFIX = '/tmp'
    XLSX = '.xlsx'

    template_fields = ('df', 'file_name', 'xlsx_params')
    ui_color = '#b4e0ff'

    def __init__(self, *,
                 df: Union[pd.DataFrame, str],
                 file_name: str,
                 xlsx_params: Optional[Union[Mapping, Iterable, dict]] = None,
                 file_exists: bool = False,
                 **kwargs) -> None:
        super().__init__(**kwargs)
        self.df = df
        self.file_name = file_name
        self.xlsx_params = xlsx_params
        self.file_exists = file_exists
        self.file = None
        self.book = None

    def pre_execute(self, context: Any):
        self.file = Path(DataFrame2Excel.PREFIX) / (self.file_name + DataFrame2Excel.XLSX)
        self.book = load_workbook(self.file) if self.file_exists else None

    def execute(self, context: Any):
        self.log.info(f"saving pandas DataFrame to file: {self.file}")

        subtotals = self.xlsx_params.pop("subtotals", None)
        group_to_sheet = self.xlsx_params.pop("group_to_sheet", None)
        title = self.xlsx_params.pop("title", None)
        self.xlsx_params["startrow"] = self.xlsx_params.get("startrow", 0) + 1 if title else self.xlsx_params \
            .get("startrow", 0)

        if group_to_sheet:
            self._split_dataframe_into_sheets(group_to_sheet, subtotals)
            if title:
                for sheet_name in self.df[group_to_sheet].unique():
                    self._add_title(title, sheet_name)
        else:
            if subtotals:
                self.df = self._add_subtotals(subtotals, self.df)

            with pd.ExcelWriter(self.file, engine='openpyxl') as writer:
                if self.file_exists:
                    writer.book = self.book
                    writer.sheets = dict((ws.title, ws) for ws in self.book.worksheets)
                self.df.to_excel(writer, **self.xlsx_params)
            if title:
                sheet_name = self.xlsx_params.pop("sheet_name", None)
                self._add_title(title, sheet_name)

        return self._file_path_as_str()

    def _split_dataframe_into_sheets(self, group_to_sheet, subtotals=None):
        with pd.ExcelWriter(self.file, engine='openpyxl') as writer:
            self._update_excel_writer(writer)

            for group, data in self.df.groupby(group_to_sheet):
                if subtotals:
                    data = self._add_subtotals(subtotals, data)
                data.to_excel(writer, str(group), **self.xlsx_params)

    def _update_excel_writer(self, writer):
        if self.file_exists:
            writer.book = self.book
            writer.sheets = dict((ws.title, ws) for ws in self.book.worksheets)

    def _file_path_as_str(self):
        return str(self.file)

    def _add_subtotals(self, subtotals, df):
        if not subtotals:
            return df

        self.log.info(f"subtotals: {subtotals}")

        if not self.xlsx_params["index"]:
            self.xlsx_params["index"] = True
        total_df = df.groupby(subtotals["group_by"]).apply(
            lambda sub_df: sub_df.pivot_table(
                index=subtotals["index"],
                values=subtotals["values"],
                aggfunc=subtotals["aggfunc"],
                margins=True,
                margins_name="TOTAL",
            )
        )

        total_df.loc[(('',) * len(subtotals["group_by"])) + ('GRAND TOTAL',)] = df[subtotals["values"]].sum()
        # preserve the order of columns
        total_df = total_df.reindex(subtotals["values"], axis=1)
        return total_df

    def _add_title(self, title, sheet_name=None):
        wb = self.book or load_workbook(self.file)
        ws = wb[sheet_name] if sheet_name else wb.active
        column = self.xlsx_params.get("startcol", 0) + 1
        row = self.xlsx_params.get("startrow", 1)
        ws.cell(column=column, row=row, value=title)
        wb.active = wb[sheet_name] if sheet_name else wb.active
        wb.save(self.file)
